import Quandl as q
from numpy import *
import matplotlib.pyplot as plt
from openpyxl.reader.excel import load_workbook
import os
import sys
import re
import pickle
from datetime import datetime
import matplotlib.ticker as ticker
import numpy as np
import pandas as pd
import math
import xlsxwriter

AUTH_TOKEN = 'e6FuWkfWH9qypKzJz6sR'
CACHE_DIR = "cache-data/"
B = sys.argv[4]
CURRENT_YEAR = datetime.now().year

def main():
    if not os.path.exists(CACHE_DIR):
        os.makedirs(CACHE_DIR)
    table = retrieveTableFromExcel()
    spread1Delta = getSpreadDelta(table[1])
    pos = spread1Delta[1][1]
    neg = spread1Delta[1][2]
    positives = spread1Delta[1][3]
    negatives = spread1Delta[1][4]
    if len(table) == 2:
        totalSpreadDelta = spread1Delta[0]
    else:
        spread2Delta = getSpreadDelta(table[2])
        pos += spread2Delta[1][1]
        neg += spread2Delta[1][2]
        positives = positives.append(spread2Delta[1][3])
        negatives = negatives.append(spread2Delta[1][4])
        totalSpreadDelta = spread1Delta[0].add(spread2Delta[0], fill_value=0)
        for i in range(0, 3):
            del table[0]
        for row in table:
            spread = getSpreadDelta(row)
            pos += spread[1][1]
            neg += spread[1][2]
            positives = positives.append(spread[1][3])
            negatives = negatives.append(spread[1][4])
            totalSpreadDelta = totalSpreadDelta.add(spread[0], fill_value=0)
    totalCumulativeChart = convertDeltaSeriesToCumulativeGraph(totalSpreadDelta)
    print("Total Cumulative Chart:")
    print(totalCumulativeChart.astype(int))
    saveReports(totalCumulativeChart, pos, neg, positives, negatives)
    showPlot(totalCumulativeChart)

def retrieveTableFromExcel():
    table = loadSpreadMatrix(sys.argv[1])
    rowsWithOneContract = []
    if (len(sys.argv) == 6):
        contract = sys.argv[5]
        rowsWithOneContract.append(table[0])
        for row in table:
            if row[0].decode('utf-8') == contract:
                rowsWithOneContract.append(row)
        table = rowsWithOneContract
    return table

def loadSpreadMatrix(filename):
    wb = load_workbook(filename)
    table = []
    for sheet in wb.worksheets:
        for row in sheet.rows:
            table_row = []
            for cell in row:
                value = cell.value
                if value is None:
                    value = ''
                if not isinstance(value, str):
                    value = str(value)
                value = value.encode('utf8')
                table_row.insert(len(table_row), value)
            table.insert(len(table), table_row)
    return table

def getSpreadDelta(row):
    if len(sys.argv) == 2:
        years = [2000]
    else:
        years = range(int(sys.argv[2]), int(sys.argv[3]) + 1)
    spread = fetchSpread(row[0].decode("utf-8"), row[1].decode("utf-8"), row[2].decode("utf-8"),
                         int(row[5][:4].decode("utf-8")), int(row[6][:4].decode("utf-8")), int(row[3].decode("utf-8")),
                         int(row[4].decode("utf-8")), row[5].decode("utf-8"), row[6].decode("utf-8"),
                         int(row[7].decode("utf-8")), True, years)
    deltaSeries = convertSpreadSeriesToDelta(spread[0])
    return (deltaSeries, spread)

def fetchSpread(CONTRACT, M1, M2, ST_YEAR, END_YEAR, CONT_YEAR1, CONT_YEAR2, ST_DATE, END_DATE, BUCK_PRICE,
                STARTFROMZERO, years):
    startdate = datetime.strptime(ST_DATE, '%Y-%m-%d %H:%M:%S')
    enddate = datetime.strptime(END_DATE, '%Y-%m-%d %H:%M:%S')
    totalSpread = pd.Series()
    lastValue = 0
    filteredTotalSpread = pd.Series()
    pos = 0
    neg = 0
    positives = pd.Series()
    negatives = pd.Series()
    for i in years:
        year = str(i)
        price = str(BUCK_PRICE)
        filename = CONTRACT + M1 + M2 + year + ST_DATE + END_DATE + price
        filename = re.sub('[/ ]', '_', filename)
        filename = re.sub('[:]', '.', filename)
        cont1 = str(CONTRACT) + str(M1) + str(i + CONT_YEAR1)
        cont2 = str(CONTRACT) + str(M2) + str(i + CONT_YEAR2)
        print('==============')
        print("contract1: " + cont1)
        print("contract2: " + cont2)
        startDate = startdate.replace(year=ST_YEAR - 2000 + i)
        endDate = enddate.replace(year=END_YEAR - 2000 + i)
        print('==============')
        print('Trim start: ', startDate.strftime('%Y-%m-%d'))
        print('Trim end: ', endDate.strftime('%Y-%m-%d'))
        print('==============')
        if not checkIfCached(filename):
            data1 = q.get(cont1, authtoken=AUTH_TOKEN, trim_start=startDate, trim_end=endDate)
            data2 = q.get(cont2, authtoken=AUTH_TOKEN, trim_start=startDate, trim_end=endDate)
            spread = (data1 - data2).Settle * BUCK_PRICE
            if spread.size == 0:
                print('!!!!!!!!!!!!*****WARNING****!!!!!!!!!!!!')
                print('No data available for contracts %s, %s. Skiping period from %s to %s.' % (
                    cont1, cont2, startDate.strftime('%Y-%m-%d'), endDate.strftime('%Y-%m-%d')))
                print('!!!!!!!!!!!!*****WARNING****!!!!!!!!!!!!')
                continue
            else:
                if math.isnan(spread[0]):
                    spread = spread.fillna(method='bfill')
                #replace NaN value with a previous one
                spread = spread.fillna(method='pad')

                #remove row with NAN value
                # spread = spread.dropna()
                writeCacheToFile(filename, spread)
        else:
            print("Loading cached data from file: %s !" % filename)
            cache = readCacheFromFile(filename)
            spread = cache
        if STARTFROMZERO:
            spreadForMeanReport = convertDeltaSeriesToCumulativeGraph(convertSpreadSeriesToDelta(spread - spread[0]))
            delta = lastValue - spread[0]
            spread = spread + delta
            if spreadForMeanReport[-1] >= 0:
                pos += 1
                positives.set_value(spreadForMeanReport.index[-1], spreadForMeanReport[-1])
            elif spreadForMeanReport[-1] < 0:
                neg += 1
                negatives.set_value(spreadForMeanReport.index[-1], spreadForMeanReport[-1])
            totalSpread = totalSpread.append(spread)
            lastValue = totalSpread[-1]
    for i in range(0, len(totalSpread)):
        if totalSpread.index[i].year <= int(sys.argv[3]):
            filteredTotalSpread.set_value(totalSpread.index[i], totalSpread[i])
    if totalSpread.size == 0:
        sys.exit(-1)
    return (filteredTotalSpread, pos, neg, positives, negatives)

def checkIfCached(filename):
    fileNames = os.listdir(CACHE_DIR)
    for fileName in fileNames:
        if fileName == filename:
            return True
    return False

def writeCacheToFile(filename, spread):
    try:
        cacheFile = open(CACHE_DIR + filename, 'wb')
        pickle.dump(spread, cacheFile)
        cacheFile.close()
    except IOError:
        print('Error: can\'t write data to %s' % (CACHE_DIR + filename))

def readCacheFromFile(filename):
    cacheFile = open(CACHE_DIR + filename, "rb")
    cache = pickle.load(cacheFile)
    cacheFile.close()
    return cache

def convertSpreadSeriesToDelta(DATA):
    DATADELTA = DATA.copy(True)
    previ = DATA.index[0]
    for i in DATA.index:
        if DATA.index[0] != i:
            DATADELTA[i] = DATA.ix[i] - DATA.ix[previ]
            previ = i
    return DATADELTA

def convertDeltaSeriesToCumulativeGraph(DATA):
    GRAPHDATA = DATA.copy(True)
    prev_date = DATA.index[0]
    for i in range(1, len(DATA.index)):
        date = DATA.index[i]
        GRAPHDATA.ix[date] = GRAPHDATA.ix[prev_date] + DATA.ix[date]
        prev_date = date
    return GRAPHDATA

def saveReports(totalCumulativeChart, pos, neg, positives, negatives):
    result = retrieveDrawdowns(totalCumulativeChart)
    print('================')
    print('Maximum drawdowns: \n', sorted(result[1], key=lambda x: x[2])[:5], '\n')
    print('================')
    yieldArray = getYieldArray(totalCumulativeChart)
    saveAllInFile(totalCumulativeChart, result[0], result[1], yieldArray, pos, neg, positives, negatives)

def retrieveDrawdowns(chart):
    res = pd.Series()
    flag = False
    prev = 0
    ind = chart.index[0]
    res.set_value(ind, prev)
    for i in range(1, len(chart)):
        if flag:
            if chart[i] > prev:
                prev = chart[i]
                ind = chart.index[i]
                res.set_value(ind, 0)
                flag = False
            else:
                ind = chart.index[i]
                res.set_value(ind, chart[i] - prev)
        else:
            if chart[i] > prev:
                prev = chart[i]
                ind = chart.index[i]
                res.set_value(ind, 0)
            else:
                ind = chart.index[i]
                res.set_value(ind, chart[i] - prev)
                flag = True
    res1 = getMAximumDDs(res)
    return (res,res1)

def getMAximumDDs(res):
    startFlag = True
    result = []
    if res[-1] != 0:
        res[-1] = 0
    for i in range(1, len(res)):
        if startFlag:
                if res[i-1] == 0 and res[i] != 0:
                    startDate = res.index[i-1]
                    startFlag = False
        else:
            if res[i] == 0 and res[i-1] != 0:
                endDate = res.index[i]
                value = min(res.truncate(startDate, endDate))
                result.append((startDate, endDate, value))
                startFlag = True
    return result

def getYieldArray(chart):
    yieldReport = []
    monthlyReport = getMonthlyReport(chart)
    yearlyReport = getYearlyReport(chart)
    dailyReport = getDailyReportWithVAMI(chart)
    yieldReport.append((monthlyReport, yearlyReport, dailyReport))
    return yieldReport

def getDailyReportWithVAMI(chart):

    def getVAMI(dailyReport):
        v1 = 1000
        vaim = [(dailyReport.index[0], v1)]
        for i in range(1 , len(dailyReport)):
            v_prev = vaim[i-1][1]
            vaim.append((dailyReport.index[i], v_prev + v1*dailyReport[i]))
        return vaim

    dailyReport = pd.Series()
    day = chart.index[0].strftime('%Y-%m-%d')
    KArray = []
    for i in range(1, len(chart)):
        if chart.index[i].strftime('%Y-%m-%d') > day:
            if chart.index[i].strftime('%Y-%m-%d') > chart.index[i - 1].strftime('%Y-%m-%d'):
                KD = chart[i - 1]
                KArray.append((chart.index[i - 1], KD))
                day = chart.index[i].strftime('%Y-%m-%d')
    firstDay = (int(KArray[0][1]) - chart[0]) / int(B)
    dailyReport.set_value(chart.index[0], firstDay)
    for j in range(1, len(KArray)):
        KD2 = KArray[j][1]
        KD1 = KArray[j - 1][1]
        yieldValue = (int(KD2) - int(KD1)) / int(B)
        dailyReport.set_value(KArray[j][0], yieldValue)
    lastDay = (chart[-1] - KD2) / int(B)
    dailyReport.set_value(chart.index[-1], lastDay)
    return (getVAMI(dailyReport), dailyReport)

def getMonthlyReport(chart):
    monthlyReport = []
    month = chart.index[0].strftime('%Y-%m')
    KArray = []
    for i in range(1, len(chart)):
        if chart.index[i].strftime('%Y-%m') > month:
            if chart.index[i].strftime('%Y-%m') > chart.index[i - 1].strftime('%Y-%m'):
                KM = chart[i - 1]
                KArray.append((chart.index[i - 1], KM))
                month = chart.index[i].strftime('%Y-%m')
    firstMonth = (int(KArray[0][1]) - chart[0]) / int(B)
    monthlyReport.append((chart.index[0], firstMonth))
    for j in range(1, len(KArray)):
        KM2 = KArray[j][1]
        KM1 = KArray[j - 1][1]
        yieldValue = (int(KM2) - int(KM1)) / int(B)
        monthlyReport.append((KArray[j][0], yieldValue))
    lastMonth = (chart[-1] - KM2) / int(B)
    monthlyReport.append((chart.index[-1], lastMonth))
    print('Monthly yield report: \n', monthlyReport)
    print('================')
    return monthlyReport

def getYearlyReport(chart):
    yearlyReport = []
    year = chart.index[0].strftime('%Y')
    KYearsArray = []
    for q in range(1, len(chart)):
        if chart.index[q].strftime('%Y') >= year:
            if chart.index[q].strftime('%Y') >= chart.index[q - 1].strftime('%Y'):
                KY = chart[q - 1]
                KYearsArray.append((chart.index[q - 1], KY))
                year = chart.index[q].strftime('%Y-%m')
    firstYear = (int(KYearsArray[0][1]) - chart[0]) / int(B)
    yearlyReport.append((chart.index[0], firstYear))
    if len(KYearsArray) == 1:
        KY2 = KYearsArray[0][1]
    for w in range(1, len(KYearsArray)):
        KY2 = KYearsArray[w][1]
        KY1 = KYearsArray[w - 1][1]
        yearYieldValue = (int(KY2) - int(KY1)) / int(B)
        yearlyReport.append((KYearsArray[w][0], yearYieldValue))
    lastYear = (chart[-1] - KY2) / int(B)
    yearlyReport.append((chart.index[-1], lastYear))
    print('Yearly yield report: \n', yearlyReport)
    print('================')
    return yearlyReport

def saveAllInFile(chart, drawdownArray, dd, yieldArray, pos, neg, positives, negatives):
    monthes = []
    for i in range(0, len(yieldArray[0][0])):
        monthes.append(yieldArray[0][0][i][1])

    workbook = xlsxwriter.Workbook('reports.xlsx')
    worksheet1 = workbook.add_worksheet("Total Cumulative Chart Report")
    worksheet2 = workbook.add_worksheet("Maximum drawdowns Report")
    worksheet3 = workbook.add_worksheet("All drawdowns Report")
    worksheet4 = workbook.add_worksheet("Monthly Report")
    worksheet5 = workbook.add_worksheet("Yearly Report")
    worksheet6 = workbook.add_worksheet("VAMI Report")
    worksheet7 = workbook.add_worksheet("Mean Values")
    worksheet8 = workbook.add_worksheet("Omega report")
    worksheet9 = workbook.add_worksheet("Daily VaR report")
    worksheet10 = workbook.add_worksheet("Monthly VaR report")

    worksheet1.set_column('A:B', 10)
    worksheet2.set_column('A:B', 10)
    worksheet3.set_column('A:B', 10)
    worksheet3.set_column('M:O', 10)
    worksheet4.set_column('A:B', 10)
    worksheet5.set_column('A:B', 10)
    worksheet6.set_column('A:B', 10)
    worksheet7.set_column('A:A', 24)
    worksheet7.set_column('B:B', 10)
    worksheet7.set_column('D:D', 19)
    worksheet7.set_column('E:E', 13)
    worksheet7.set_column('G:G', 19)
    worksheet7.set_column('H:H', 15)
    worksheet8.set_column('A:B', 10)
    worksheet9.set_column('A:H', 10)
    worksheet10.set_column('A:H', 10)
    chart1 = getTCCChart(workbook, worksheet1, chart)
    chart2 = getChartWithMaximumDrowdowns(workbook, worksheet2, dd)
    chart3 = getChartWithAllDrawdowns(workbook, worksheet3, drawdownArray, dd)
    chart4 = getMonthlyChart(workbook, worksheet4, yieldArray)
    chart5 = getYearlyChart(workbook, worksheet5, yieldArray)
    chart6 = getDailyChart(workbook, worksheet6, yieldArray)
    chart8 = getOmegaChart(workbook, worksheet8, yieldArray[0][0])
    chart9 = getVaRChart(workbook, worksheet9, yieldArray[0][2][1], 'Daily VaR report')
    chart10 = getVaRChart(workbook, worksheet10, monthes, 'Monthly VaR report')
    worksheet1.insert_chart('C1', chart1)
    worksheet2.insert_chart('E1', chart2)
    worksheet3.insert_chart('C1', chart3)
    worksheet4.insert_chart('C1', chart4)
    worksheet5.insert_chart('C1', chart5)
    worksheet6.insert_chart('C1', chart6)
    worksheet8.insert_chart('D3', chart8)
    worksheet9.insert_chart('A12', chart9)
    worksheet10.insert_chart('A12', chart10)

    # merge_format = workbook.add_format({'align': 'center'})
    # worksheet7.merge_range('A1:B1', 'Monthly mean', merge_format)
    # worksheet7.merge_range('D1:E1', 'Yearly mean', merge_format)
    worksheet7.write_string(0, 0, 'Count of positive deals:')
    worksheet7.write_string(1, 0, 'Count of negative deals:')
    worksheet7.write_string(2, 0, 'Mean of positive deals:')
    worksheet7.write_string(3, 0, 'Mean of negative deals:')
    worksheet7.write_string(4, 0, 'Percentage of positive deals:')
    worksheet7.write_string(5, 0, 'Mean of all deals:')
    worksheet7.write_string(6, 0, 'Total profit:')
    worksheet7.write_string(7, 0, 'Total trades:')
    worksheet7.write_string(8, 0, 'Ratio:')
    worksheet7.write_string(9, 0, 'Gross profit:')
    worksheet7.write_string(10, 0, 'Gross loss:')
    worksheet7.write_string(11, 0, 'Profit factor:')

    worksheet7.write_string(0, 3, 'Daily stdev:')
    worksheet7.write_string(1, 3, 'Monthly stdev:')
    worksheet7.write_string(2, 3, 'Yearly stdev:')
    worksheet7.write_string(3, 3, 'Average daily yield:')
    worksheet7.write_string(4, 3, 'Average monthly yield:')
    worksheet7.write_string(5, 3, 'Average yearly yield:')
    worksheet7.write_string(6, 3, 'Daily average gain:')
    worksheet7.write_string(7, 3, 'Monthly average gain:')
    worksheet7.write_string(8, 3, 'Yearly average gain:')
    worksheet7.write_string(9, 3, 'Daily average loss:')
    worksheet7.write_string(10, 3, 'Monthly average loss:')
    worksheet7.write_string(11, 3, 'Yearly average loss:')
    worksheet7.write_string(12, 3, 'Daily stdev gain:')
    worksheet7.write_string(13, 3, 'Monthly stdev gain:')
    worksheet7.write_string(14, 3, 'Yearly stdev gain:')
    worksheet7.write_string(15, 3, 'Daily stdev loss:')
    worksheet7.write_string(16, 3, 'Monthly stdev loss:')
    worksheet7.write_string(17, 3, 'Yearly stdev loss:')
    worksheet7.write_string(18, 3, 'Daily avg gain/avg:')
    worksheet7.write_string(19, 3, 'Monthly avg gain/avg:')
    worksheet7.write_string(20, 3, 'Yearly avg gain/avg:')

    worksheet7.write_string(0, 6, 'Calmar Ratio:')
    worksheet7.write_string(1, 6, 'Sterling Ratio:')
    worksheet7.write_string(2, 6, 'Sharpe Ratio monthly:')
    worksheet7.write_string(3, 6, 'Sharpe Ratio yearly:')
    worksheet7.write_string(4, 6, 'Sharpe Ratio2 daily:')
    worksheet7.write_string(5, 6, 'Sharpe Ratio2 monthly:')
    worksheet7.write_string(6, 6, 'Sharpe Ratio2 yearly:')
    worksheet7.write_string(7, 6, 'Sortino ratio:')
    worksheet7.write_string(8, 6, 'Downside deviation:')
    worksheet7.write_string(9, 6, 'Daily skewness:')
    worksheet7.write_string(10, 6, 'Monthly skewness:')
    worksheet7.write_string(11, 6, 'Yearly skewness:')
    worksheet7.write_string(12, 6, 'Daily kurtosis:')
    worksheet7.write_string(13, 6, 'Monthly kurtosis:')
    worksheet7.write_string(14, 6, 'Yearly kurtosis:')
    writeTransactionsAmount(positives, negatives, worksheet7, workbook, pos, neg, chart, yieldArray, dd)

    workbook.close()

def getTCCChart(workbook, worksheet, tcc):
    chart = workbook.add_chart({'type': 'line'})
    a = 0
    b = 0
    row = 0
    col = 0
    for index in (tcc.index):
        date = datetime.strftime(index, '%Y-%m-%d')
        worksheet.write_string(row, col, date)
        a += 1
        row += 1
    row = 0
    for value in (tcc):
        worksheet.write_number(row, col + 1, int(value))
        b += 1
        row += 1
    chart.set_x_axis({
        'date_axis': True
    })
    chart.add_series({
        'values': '=Total Cumulative Chart Report!$B$1:$B$' + str(b),
        'categories': '=Total Cumulative Chart Report!$A$1:$A$' + str(a)
    })
    chart.set_y_axis({
        'major_gridlines': {
            'visible': True,
            'line': {'width': 1.25, 'dash_type': 'dash'}
        }
    })
    chart.set_size({'width': 720, 'height': 570})
    return chart

def getChartWithMaximumDrowdowns(workbook, worksheet, dd):
    firstDate = []
    secondDate = []
    delta = []
    sortedDDArray = sorted(dd, key=lambda x: x[2])[:5]
    chart = workbook.add_chart({'type': 'column'})
    for i in range(0, len(sortedDDArray)):
        firstDate.append(sortedDDArray[i][0])
        secondDate.append(sortedDDArray[i][1])
        delta.append(sortedDDArray[i][2])
    col = 0
    row = 0
    for date1 in firstDate:
        dateLeft = datetime.strftime(date1, '%Y-%m-%d')
        worksheet.write_string(row, col, dateLeft)
        row += 1
    row = 0
    for date2 in secondDate:
        dateRight = datetime.strftime(date2, '%Y-%m-%d')
        worksheet.write_string(row, col + 1, dateRight)
        row += 1
    row = 0
    for d in delta:
        worksheet.write_number(row, col + 2, int(d))
        worksheet.write_number(row, col + 3, int(d)/int(B))
        row += 1
    chart.add_series({
        'values': '=Maximum drawdowns Report!$C$1:$C$5',
        'categories': '=Maximum drawdowns Report!$A$1:$A$5'
    })
    chart.set_size({'width': 600, 'height': 470})
    return chart

def getChartWithAllDrawdowns(workbook, worksheet, sortedDDArray, dd):
    chart = workbook.add_chart({'type': 'area'})
    firstDate = []
    secondDate = []
    delta = []
    for i in range(0, len(dd)):
        firstDate.append(dd[i][0])
        secondDate.append(dd[i][1])
        delta.append(dd[i][2])
    writeArrayOFAllDDs(firstDate, secondDate, delta, worksheet)
    col = 0
    row = 0
    a = 0
    c = 0
    for date1 in (sortedDDArray.index):
        dateLeft = datetime.strftime(date1, '%Y-%m-%d')
        worksheet.write_string(row, col, dateLeft)
        row += 1
        a += 1
    row = 0
    for d in (sortedDDArray):
        worksheet.write_number(row, col + 1, int(d))
        row += 1
        c += 1
    chart.add_series({
        'values': '=All drawdowns Report!$B$1:$B$' + str(c),
        'categories': '=All drawdowns Report!$A$1:$A$' + str(a)
    })
    chart.set_size({'width': 600, 'height': 470})
    return chart

def writeArrayOFAllDDs(firstDate, secondDate, delta, worksheet):
    row = 0
    col = 12

    for date1 in firstDate:
        dateLeft = datetime.strftime(date1, '%Y-%m-%d')
        worksheet.write_string(row, col, dateLeft)
        row += 1
    row = 0
    for date2 in secondDate:
        dateRight = datetime.strftime(date2, '%Y-%m-%d')
        worksheet.write_string(row, col+1, dateRight)
        row += 1
    row = 0
    for d in delta:
        worksheet.write_number(row, col + 2, int(d))
        row += 1

def getMonthlyChart(workbook, worksheet, yieldArray):
    m_dates = []
    m_values = []
    chart = workbook.add_chart({'type': 'column'})
    for i in range(0, len(yieldArray[0][0])):
        m_dates.append(yieldArray[0][0][i][0])
        m_values.append(yieldArray[0][0][i][1])
    col = 0
    row = 0
    am = 0
    bm = 0
    for date in m_dates:
        worksheet.write_string(row, col, datetime.strftime(date, '%Y-%m-%d'))
        row += 1
        am += 1
    row = 0
    for value in m_values:
        worksheet.write_number(row, col + 1, value)
        row += 1
        bm += 1
    worksheet.write_string(3, 14, "MEAN:")
    worksheet.write_number(3, 15, (sum(m_values)/len(m_values)))
    chart.add_series({
        'values': '=Monthly Report!$B$1:$B$' + str(bm),
        'categories': '=Monthly Report!$A$1:$A$' + str(am)
    })
    chart.set_size({'width': 720, 'height': 570})
    return chart

def getYearlyChart(workbook, worksheet, yieldArray):
    y_dates = []
    y_values = []
    chart = workbook.add_chart({'type': 'column'})
    for j in range(0, len(yieldArray[0][1])):
        y_dates.append(yieldArray[0][1][j][0])
        y_values.append(yieldArray[0][1][j][1])
    col = 0
    row = 0
    a = 0
    b = 0
    for date in y_dates:
        worksheet.write_string(row, col, datetime.strftime(date, '%Y-%m-%d'))
        row += 1
        a += 1
    row = 0
    for value in y_values:
        worksheet.write_number(row, col + 1, value)
        row += 1
        b += 1
    worksheet.write_string(3, 14, "MEAN:")
    worksheet.write_number(3, 15, (sum(y_values)/len(y_values)))
    chart.add_series({
        'values': '=Yearly Report!$B$1:$B$' + str(b),
        'categories': '=Yearly Report!$A$1:$A$' + str(a)
    })
    chart.set_size({'width': 720, 'height': 570})
    return chart

def getDailyChart(workbook, worksheet, yieldArray):
    d_dates = []
    d_values = []
    chart = workbook.add_chart({'type': 'column'})
    for i in range(0,len(yieldArray[0][2][0])):
        d_dates.append(yieldArray[0][2][0][i][0])
        d_values.append(yieldArray[0][2][0][i][1])
    col = 0
    row = 0
    a = 0
    b = 0
    for date in d_dates:
        worksheet.write_string(row, col, datetime.strftime(date, '%Y-%m-%d'))
        row += 1
        a += 1
    row = 0
    for value in d_values:
        worksheet.write_number(row, col + 1, value)
        row += 1
        b += 1
    chart.add_series({
        'values': '=VAMI Report!$B$1:$B$' + str(b),
        'categories': '=VAMI Report!$A$1:$A$' + str(a)
    })
    chart.set_size({'width': 720, 'height': 570})
    return chart

def getOmegaChart(workbook, worksheet, monthlyYield):

    def getOmega(const):
        marArr = []
        for m in range(0, len(monthlyYield)):
            marArr.append(monthlyYield[m][1]-const)
        return getOmegaValue(marArr)

    def getOmegaValue(arr):
        sumPos = 0
        sumNeg = 0
        for a in arr:
            if a >= 0:
                sumPos += a
            else:
                sumNeg += a
        return 0 if sumNeg == 0 else sumPos/abs(sumNeg)

    format = workbook.add_format()
    format.set_num_format('0.00%')
    omega_chart = workbook.add_chart({'type': 'line'})
    worksheet.write_string(0, 0, 'MAR, year')
    worksheet.write_string(0, 1, 'MAR, month')
    worksheet.write_string(0, 3, 'MAR')
    worksheet.write_string(1, 3, 'Omega')
    marConst = []
    row = 1
    col = 0
    row1 = 0
    col1 = 4
    mar = 0.12
    for i in range (1, 11):
        worksheet.write_number(row, col, mar, format)
        worksheet.write_number(row, col+1, mar/12, format)
        marConst.append(round(mar/12, 4))
        worksheet.write_number(row1, col1, mar)
        row += 1
        col1 += 1
        mar += 0.02
    row = 1
    col = 4
    for q in range(0, 10):
        worksheet.write_number(row, col, getOmega(marConst[q]))
        col += 1

    omega_chart.add_series({
        'values': '=Omega report!$E$2:$N$2',
        'categories': '=Omega report!$E$1:$N$1'
    })

    omega_chart.set_size({'width': 520, 'height': 370})

    return omega_chart

def getVaRChart(workbook, worksheet, yieldArray, name):
    def stdev(x):
       return sqrt(sum((x - mean(x))**2)/(len(x)-1)) if len(x) > 1 else sqrt(sum((x - mean(x))**2)/len(x))

    chart = workbook.add_chart({'type':'line'})
    format = workbook.add_format()
    format.set_num_format('0.00%')
    length = len(yieldArray)
    worksheet.write_string(0, 0, 'Confidence,%')
    worksheet.write_string(0, 1, 'Z(quantile)')
    worksheet.write_string(0, 2, 'VaR,normal')
    worksheet.write_string(0, 3, 'Z,modified')
    worksheet.write_string(0, 4, 'VaR,modified')
    worksheet.write_string(0, 5, 'Bott-1')
    worksheet.write_string(0, 6, 'Bott-2')
    worksheet.write_string(0, 7, 'VaR,historical')

    row = 1
    col = 0
    conf = 0.9
    confidence = []
    for q in range(0, 10):
        worksheet.write_number(row, col, conf)
        confidence.append(conf)
        conf += 0.01
        row += 1

    quantile = [-1.28155157, -1.34075503, -1.40507156, -1.47579103, -1.55477359, -1.64485363, -1.75068607, -1.88079361, -2.05374891, -2.32634787]

    col = 1
    row = 1
    for i in range(0, 10):
        worksheet.write_number(row+i, col, quantile[i])
        i += 1

    avgDailyYield = mean(yieldArray)
    dailyStdev = stdev(yieldArray)
    normalVAR = []
    for k in range(0, 10):
        normalVAR.append(avgDailyYield + (quantile[k]*dailyStdev))

    col = 2
    row = 1
    for w in range(0, 10):
        worksheet.write_number(row+w, col, normalVAR[w], format)
        w += 1

    row = 1
    col  = 5
    for e in range(0, len(normalVAR)):
        koef = length*(1-confidence[e])
        ind = koef if koef > 1 else 1
        sortedYield = sorted(yieldArray)
        bott1 = sortedYield[round(ind)-1]
        bott2 = sortedYield[int(ind)]

        historicalVar = bott1 + (bott2 - bott1)*(koef - int(koef))
        worksheet.write_number(row+e, col, bott1)
        worksheet.write_number(row+e, col+1, bott2)
        worksheet.write_number(row+e, col+2, historicalVar, format)

    worksheet.write_string(0, 9, 'Yield array')
    row = 1
    col = 9
    for value in yieldArray:
        worksheet.write_number(row, col, value)
        row += 1

    chart.add_series({
        'name':   '='+name+'!$C$1',
        'values' : '='+name+'!$C$2:$C$11',
        'categories' : '='+name+'!$A$2:$A$11'
    })
    chart.add_series({
        'name':   '='+name+'!$H$1',
        'values' : '='+name+'!$H$2:$H$11',
        'categories' : '='+name+'!$A$2:$A$11'
    })
    chart.set_size({'width': 620, 'height': 370})

    return chart

def getMonthlyVaRChart(workbook, worksheet10, monthlyYield):
    pass

# def getDistributionChart(workbook, worksheet, monthlyYield):
#     distribution_chart = workbook.add_chart({'type': 'line'})
#     hist_chart = workbook.add_chart({'type': 'bar'})
#
#     worksheet.write_string(0,0,'interval')
#     worksheet.write_string(1,0,'norm distrib')
#     worksheet.write_string(0,0,'frequency')
#     format = workbook.add_format()
#     format.set_num_format('0.00')
#     a = 0
#     b = 0
#     c = 0
#     col = 0
#     row = 1
#     # for q in x:
#     #     worksheet.write_number(row, col, q)
#     #     row += 1
#     #     a += 1
#     # row = 1
#     # for y in norm_distribution:
#     #     worksheet.write_number(row, col+1, y)
#     #     row += 1
#     #     b += 1
#     # row = 1
#     # for o in hist[0]:
#     #     worksheet.write_number(row, col+2, o)
#     #     row += 1
#     #     c += 1
#
#     distribution_chart.add_series({
#         'values': '=Norm Distribution Report!$B$1:$B$' + str(b),
#         'categories': '=Norm Distribution Report!$A$1:$A$' + str(a)
#     })
#     hist_chart.add_series({
#         'values': '=Norm Distribution Report!$A$1:$A$' + str(a),
#         'categories': '=Norm Distribution Report!$C$1:$C$' + str(c)
#     })
#     distribution_chart.set_size({'height': 570})
#     hist_chart.set_size({'height': 570})
#     return (distribution_chart, hist_chart)

def writeTransactionsAmount(positiveSeries, negativeSeries, w_sheet, w_book, pos, neg, chart, yieldArray, dd):
    negativeValue = 0
    positiveValue = 0
    positiveMean = 0
    negativeMean = 0
    for i in range(0, len(positiveSeries)):
        positiveValue += positiveSeries[i]
    for j in range(0, len(negativeSeries)):
        negativeValue += negativeSeries[j]

    format = w_book.add_format()
    format.set_num_format('0.00%')
    format1 = w_book.add_format()
    format1.set_num_format('0.00')
    w_sheet.write_number(0, 1, pos)#count of positive deals
    w_sheet.write_number(1, 1, neg)#count of negative deals
    #mean of positives
    if len(positiveSeries) == 0:
        w_sheet.write_number(2, 1, positiveMean, format1)
    else:
        positiveMean = mean(positiveSeries)
        w_sheet.write_number(2, 1, positiveMean, format1)
    #mean of negatives
    if len(negativeSeries) == 0:
        w_sheet.write_number(3, 1, negativeMean, format1)
    else:
        negativeMean =  mean(negativeSeries)
        w_sheet.write_number(3, 1, negativeMean, format1)
    w_sheet.write_number(4, 1, (pos/(len(positiveSeries) + len(negativeSeries))), format)#percentage of positives
    w_sheet.write_number(5, 1, (positiveValue + negativeValue)/(len(positiveSeries) + len(negativeSeries)), format1)#mean of all deals
    w_sheet.write_number(6, 1, chart[-1], format1)#total profit
    w_sheet.write_number(7, 1, pos + neg)#total trades
    #ratio
    if negativeMean == 0:
        w_sheet.write_number(8, 1, 0)
    else:
        w_sheet.write_number(8, 1, abs(positiveMean/negativeMean), format1)
    w_sheet.write_number(9, 1, positiveValue, format1)#gross profit
    w_sheet.write_number(10, 1, negativeValue, format1)#gross loss
    #profit factor
    if negativeValue == 0:
        w_sheet.write_number(11, 1, 0)
    else:
        w_sheet.write_number(11, 1, abs(positiveValue/negativeValue), format1)
    calculateAvgInYield(yieldArray[0][2][1], yieldArray[0][0], yieldArray[0][1], w_sheet, dd)

def calculateAvgInYield(dailyYield, monthlyYield, yearlyYield, w_sheet, dd):
    def stdev(x):
       return sqrt(sum((x - mean(x))**2)/(len(x)-1)) if len(x) > 1 else sqrt(sum((x - mean(x))**2)/len(x))

    def skew(x, avg, stdev):
        n = len(x)
        arr = ((x - avg)/stdev)**3
        summary = sum(arr)
        return n*summary/((n-1)*(n-2)) if n > 2 else 0

    def kurtosis(x, avg, stdev):
        n = len(x)
        arr = ((x - avg)/stdev)**4
        summary = sum(arr)
        a = (n*(n + 1))/((n-1)*(n-2)*(n-3)) if n > 3 else 0
        b = 3*(n-1)**2/((n-2)*(n-3)) if n > 3 else 0
        return a*summary - b

    def calcMean(arr):
        positiveSum = negativeSum = 0
        countPos = countNeg = 0
        for x in arr:

            if x >= 0:
                positiveSum += x
                countPos += 1
            else:
                negativeSum += x
                countNeg += 1
        meanPos = 0 if countPos == 0 else positiveSum/countPos
        meanNeg = 0 if countNeg == 0 else negativeSum/countNeg
        return (meanPos, meanNeg)

    def calcStdev(arr):
        posArr = []
        negArr = []
        for x in arr:
            if x >= 0:
                posArr.append(x)
            else:
                negArr.append(x)
        posStdev = stdev(posArr) if len(posArr) > 0 else 0
        negStdev = stdev(negArr) if len(negArr) > 0 else 0
        return (posStdev, negStdev)

    def calcRatio(pos, neg):
        return 0 if neg == 0 else (pos/neg)

    def downside_dev(arr):
        result = []
        for i in arr:
            val = i - 0.0167
            if val >= 0:
               result.append(0)
            else:
               result.append(val)
        return stdev(result)

    monthes=[]
    for b in range(0, len(monthlyYield)):
        monthes.append(monthlyYield[b][1])
    years=[]
    for c in range(0, len(yearlyYield)):
        years.append(yearlyYield[c][1])
    dailyStdev =  stdev(dailyYield)
    monthlyStdev = stdev(monthes)
    yearlyStdev = stdev(years)
    avgDailyYield = mean(dailyYield)
    avgMonthlyYield = mean(monthes)
    avgYearlyYield = mean(years)
    w_sheet.write_number(0, 4, dailyStdev)#daily stdev
    w_sheet.write_number(1, 4, monthlyStdev)#monthly stdev
    w_sheet.write_number(2, 4, yearlyStdev)#yearly stdev
    w_sheet.write_number(3, 4, avgDailyYield)#average daily yield
    w_sheet.write_number(4, 4, avgMonthlyYield)#average monthly yield
    w_sheet.write_number(5, 4, avgYearlyYield)#average yearly yield

    #Daily average gain, Monthly average gain, Yearly average gain, Daily average loss, Monthly average loss, Yearly average loss
    dayPosMean, dayNegMean = calcMean(dailyYield)
    monthPosMean, monthNegMean = calcMean(monthes)
    yearPosMean, yearNegMean = calcMean(years)
    w_sheet.write_number(6, 4, dayPosMean)
    w_sheet.write_number(7, 4, monthPosMean)
    w_sheet.write_number(8, 4, yearPosMean)
    w_sheet.write_number(9, 4, dayNegMean)
    w_sheet.write_number(10, 4, monthNegMean)
    w_sheet.write_number(11, 4, yearNegMean)

    #Standard deviations gain and loss
    dayPosStdev, dayNegStdev = calcStdev(dailyYield)
    monthPosStdev, monthNegStdev = calcStdev(monthes)
    yearPosStdev, yearNegStdev = calcStdev(years)
    w_sheet.write_number(12, 4, dayPosStdev)
    w_sheet.write_number(13, 4, monthPosStdev)
    w_sheet.write_number(14, 4, yearPosStdev)
    w_sheet.write_number(15, 4, dayNegStdev)
    w_sheet.write_number(16, 4, monthNegStdev)
    w_sheet.write_number(17, 4, yearNegStdev)

    #avg gain/avg loss
    dayRatio = calcRatio(dayPosMean, dayNegMean)
    monthRatio = calcRatio(monthPosMean, monthNegMean)
    yearRatio = calcRatio(yearPosMean, yearNegMean)
    w_sheet.write_number(18, 4, dayRatio)
    w_sheet.write_number(19, 4, monthRatio)
    w_sheet.write_number(20, 4, yearRatio)

    #Calmar ratio, Sterling ratio, Sharpe ratios, Sortino ratio
    maxDD = min(dd,key=lambda x:x[2])[2]/int(B)
    calmar = avgYearlyYield/abs(maxDD)
    sterling = avgYearlyYield/abs(maxDD - 0.1)
    sharpeRatioYearly = (avgYearlyYield - 0.05)/yearlyStdev
    sharpeRatioMonthly = (avgMonthlyYield - 0.0042)/monthlyStdev
    sharpRatio2Daily = sqrt(253)*avgDailyYield/dailyStdev
    sharpRatio2Monthly = sqrt(12)*avgMonthlyYield/monthlyStdev
    sharpRatio2Yearly = avgYearlyYield/yearlyStdev
    downSideDev = downside_dev(monthes)
    sortinoRatio = (avgMonthlyYield - 0.0167)/downSideDev if downSideDev != 0 else 0
    dailySkew = skew(dailyYield, avgDailyYield, dailyStdev)
    dailyKurtosis = kurtosis(dailyYield, avgDailyYield, dailyStdev)
    monthlySkew = skew(monthes, avgMonthlyYield, monthlyStdev)
    monthlyKurtosis = kurtosis(monthes, avgMonthlyYield, monthlyStdev)
    yearlySkew = skew(years, avgYearlyYield, yearlyStdev)
    yearlyKurtosis = kurtosis(years, avgYearlyYield, yearlyStdev)
    w_sheet.write_number(0, 7, calmar)
    w_sheet.write_number(1, 7, sterling)
    w_sheet.write_number(2, 7, sharpeRatioMonthly)
    w_sheet.write_number(3, 7, sharpeRatioYearly)
    w_sheet.write_number(4, 7, sharpRatio2Daily)
    w_sheet.write_number(5, 7, sharpRatio2Monthly)
    w_sheet.write_number(6, 7, sharpRatio2Yearly)
    w_sheet.write_number(7, 7, sortinoRatio)
    w_sheet.write_number(8, 7, downSideDev)
    w_sheet.write_number(9, 7, dailySkew)
    w_sheet.write_number(10, 7, monthlySkew)
    w_sheet.write_number(11, 7, yearlySkew)
    w_sheet.write_number(12, 7, dailyKurtosis)
    w_sheet.write_number(13, 7, monthlyKurtosis)
    w_sheet.write_number(14, 7, yearlyKurtosis)

def showPlot(totalCumulativeChart):
    def format_date(x, pos=None):
        thisind = np.clip(int(x + 0.5), 0, N - 1)
        return totalCumulativeChart.index[thisind].strftime('%b %d %Y')

    N = len(totalCumulativeChart)
    ind = np.arange(N)

    #shows plot with empty data intervals
    # fig, ax = plt.subplots()
    # ax.plot(totalCumulativeChart.index, totalCumulativeChart)
    # fig.autofmt_xdate()

    #shows plot without empty data intervals
    fig, ax = plt.subplots()
    ax.plot(ind, totalCumulativeChart)
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_date))
    fig.autofmt_xdate()

    ax.yaxis.grid()
    plt.xticks(np.arange(min(ind), max(ind), 15))
    plt.show()



main()
