import Quandl as q
from numpy import *
import matplotlib.pyplot as plt
from openpyxl.reader.excel import load_workbook
import os
import sys
import re
import pickle
from datetime import datetime
import matplotlib.ticker as ticker
import numpy as np
import pandas as pd
import math
import xlsxwriter

AUTH_TOKEN = 'e6FuWkfWH9qypKzJz6sR'
CACHE_DIR = "cache-data/"
CURRENT_YEAR = datetime.now().year

def main():
    if not os.path.exists(CACHE_DIR):
        os.makedirs(CACHE_DIR)
    unProf = []
    table = retrieveTableFromExcel()
    spread1Delta = getSpreadDelta(table[1])
    unProf.append(spread1Delta)
    if len(table) == 2:
        totalSpreadDelta = spread1Delta[0]
    else:
        spread2Delta = getSpreadDelta(table[2])
        unProf.append(spread2Delta)
        totalSpreadDelta = spread1Delta[0].add(spread2Delta[0], fill_value=0)
        for i in range(0, 3):
            del table[0]
        for row in table:
            spread = getSpreadDelta(row)
            unProf.append(spread)
            totalSpreadDelta = totalSpreadDelta.add(spread[0], fill_value=0)
    totalCumulativeChart = convertDeltaSeriesToCumulativeGraph(totalSpreadDelta)
    saveReports(unProf)
    print("Total Cumulative Chart:")
    print(totalCumulativeChart.astype(int))
    showPlot(totalCumulativeChart)

def retrieveTableFromExcel():
    table = loadSpreadMatrix(sys.argv[1])
    rowsWithOneContract = []
    if (len(sys.argv) == 5):
        contract = sys.argv[4]
        rowsWithOneContract.append(table[0])
        for row in table:
            if row[0].decode('utf-8') == contract:
                rowsWithOneContract.append(row)
        table = rowsWithOneContract
    return table

def loadSpreadMatrix(filename):
    wb = load_workbook(filename)
    table = []
    for sheet in wb.worksheets:
        for row in sheet.rows:
            table_row = []
            for cell in row:
                value = cell.value
                if value is None:
                    value = ''
                if not isinstance(value, str):
                    value = str(value)
                value = value.encode('utf8')
                table_row.insert(len(table_row), value)
            table.insert(len(table), table_row)
    return table

def getSpreadDelta(row):
    if len(sys.argv) == 2:
        years = [2000]
    else:
        years = range(int(sys.argv[2]), int(sys.argv[3]) + 1)

    coeft = 0 if row[10].decode('utf-8') == '' else int(row[10].decode('utf-8'))
    spread = fetchSpread(row[0].decode("utf-8"), row[1].decode("utf-8"), row[2].decode("utf-8"),
                         int(row[5][:4].decode("utf-8")), int(row[6][:4].decode("utf-8")), int(row[3].decode("utf-8")),
                         int(row[4].decode("utf-8")), row[5].decode("utf-8"), row[6].decode("utf-8"),
                         int(row[7].decode("utf-8")), int(row[8].decode('utf-8')), coeft, True, years)
    deltaSeries = convertSpreadSeriesToDelta(spread[0])
    return (deltaSeries, spread)

def fetchSpread(CONTRACT, M1, M2, ST_YEAR, END_YEAR, CONT_YEAR1, CONT_YEAR2, ST_DATE, END_DATE, BUCK_PRICE, COMISSION,
                COEFT,
                STARTFROMZERO, years):
    startdate = datetime.strptime(ST_DATE, '%Y-%m-%d %H:%M:%S')
    enddate = datetime.strptime(END_DATE, '%Y-%m-%d %H:%M:%S')
    totalSpread = pd.Series()
    lastValue = 0
    filteredTotalSpread = pd.Series()
    pos = 0
    neg = 0
    positives = pd.Series()
    negatives = pd.Series()
    spreadForStdevs = []
    for i in years:
        year = str(i)
        price = str(BUCK_PRICE)
        filename = CONTRACT + M1 + M2 + year + ST_DATE + END_DATE + price
        filename = re.sub('[/ ]', '_', filename)
        filename = re.sub('[:]', '.', filename)
        cont1 = str(CONTRACT) + str(M1) + str(i + CONT_YEAR1)
        cont2 = str(CONTRACT) + str(M2) + str(i + CONT_YEAR2)
        print('==============')
        print("contract1: " + cont1)
        print("contract2: " + cont2)
        startDate = startdate.replace(year=ST_YEAR - 2000 + i)
        endDate = enddate.replace(year=END_YEAR - 2000 + i)
        print('==============')
        print('Trim start: ', startDate.strftime('%Y-%m-%d'))
        print('Trim end: ', endDate.strftime('%Y-%m-%d'))
        print('==============')
        if not checkIfCached(filename):
            data1 = q.get(cont1, authtoken=AUTH_TOKEN, trim_start=startDate, trim_end=endDate)
            data2 = q.get(cont2, authtoken=AUTH_TOKEN, trim_start=startDate, trim_end=endDate)
            spread = (data1 - data2).Settle * BUCK_PRICE * COEFT - (COMISSION * COEFT) if COEFT != 0 else (
                                                                                                              data1 - data2).Settle * BUCK_PRICE - COMISSION
            if spread.size == 0:
                print('!!!!!!!!!!!!*****WARNING****!!!!!!!!!!!!')
                print('No data available for contracts %s, %s. Skiping period from %s to %s.' % (
                    cont1, cont2, startDate.strftime('%Y-%m-%d'), endDate.strftime('%Y-%m-%d')))
                print('!!!!!!!!!!!!*****WARNING****!!!!!!!!!!!!')
                continue
            else:
                if math.isnan(spread[0]):
                    spread = spread.fillna(method='bfill')
                #replace NaN value with a previous one
                spread = spread.fillna(method='pad')

                #remove row with NAN value
                # spread = spread.dropna()
                writeCacheToFile(filename, spread)
        else:
            print("Loading cached data from file: %s !" % filename)
            cache = readCacheFromFile(filename)
            spread = cache * COEFT if COEFT != 0 else cache
        if STARTFROMZERO:
            spreadForMeanReport = convertDeltaSeriesToCumulativeGraph(convertSpreadSeriesToDelta(spread - spread[0]))
            delta = lastValue - spread[0]
            spread = spread + delta - COMISSION
            if spreadForMeanReport[-1] >= 0:
                pos += 1
                positives.set_value(spreadForMeanReport.index[-1], spreadForMeanReport[-1])
            elif spreadForMeanReport[-1] < 0:
                neg += 1
                negatives.set_value(spreadForMeanReport.index[-1], spreadForMeanReport[-1])
            spreadForStdevs.append(spreadForMeanReport[-1])
            totalSpread = totalSpread.append(spread)
            lastValue = totalSpread[-1]
    for i in range(0, len(totalSpread)):
        if totalSpread.index[i].year <= int(sys.argv[3]):
            filteredTotalSpread.set_value(totalSpread.index[i], totalSpread[i])
    if totalSpread.size == 0:
        sys.exit(-1)
    return (filteredTotalSpread, pos, neg, positives, negatives)

def checkIfCached(filename):
    fileNames = os.listdir(CACHE_DIR)
    for fileName in fileNames:
        if fileName == filename:
            return True
    return False

def writeCacheToFile(filename, spread):
    try:
        cacheFile = open(CACHE_DIR + filename, 'wb')
        pickle.dump(spread, cacheFile)
        cacheFile.close()
    except IOError:
        print('Error: can\'t write data to %s' % (CACHE_DIR + filename))

def readCacheFromFile(filename):
    cacheFile = open(CACHE_DIR + filename, "rb")
    cache = pickle.load(cacheFile)
    cacheFile.close()
    return cache

def convertSpreadSeriesToDelta(DATA):
    DATADELTA = DATA.copy(True)
    previ = DATA.index[0]
    for i in DATA.index:
        if DATA.index[0] != i:
            DATADELTA[i] = DATA.ix[i] - DATA.ix[previ]
            previ = i
    return DATADELTA

def convertDeltaSeriesToCumulativeGraph(DATA):
    GRAPHDATA = DATA.copy(True)
    prev_date = DATA.index[0]
    for i in range(1, len(DATA.index)):
        date = DATA.index[i]
        GRAPHDATA.ix[date] = GRAPHDATA.ix[prev_date] + DATA.ix[date]
        prev_date = date
    return GRAPHDATA

def getMarginTuple(row):
    start = row[5].decode('utf-8')
    end = row[6].decode('utf-8')
    margin = int(row[9].decode('utf-8'))
    return ((start, end, margin))

def saveReports(profits):
    workbook = xlsxwriter.Workbook('unrealized profit.xlsx')
    worksheet = workbook.add_worksheet("Unrealized profit")
    worksheet.set_column('A:A', 10)
    chart = getMarginChart(workbook, worksheet, profits)
    worksheet.insert_chart('C1', chart)
    workbook.close()

def getMarginChart(workbook, worksheet, margins):
    print("Starting to write unrealized profit report at ", datetime.now())
    result = []
    chart = workbook.add_chart({'type': 'line'})
    format = '%Y-%m-%d'
    profitYears = pd.date_range(datetime(margins[0][0].index[0].year, 1, 1),
                                datetime(margins[0][0].index[-1].year + 1, 1, 1))
    for i in range(0, len(profitYears)):
        profit = 0
        date = datetime.strftime(profitYears[i], format)
        for j in range(0, len(margins)):
            spread = margins[j][0]
            for s in range(0, len(spread)):
                if str(spread.index[s]) <= date <= str(spread.index[-1]):
                    try:
                        profit += margins[j][0][date]
                        # print(profit)
                        # s+=1
                        break
                    except KeyError:
                        # print("error in method ", inspect.stack()[0][3])
                        break

        result.append((profitYears[i], profit))

    row = 0
    col = 0
    a = 0
    for r in result:
        date = datetime.strftime(r[0], '%Y-%m-%d')
        worksheet.write_string(row, col, date)
        worksheet.write_number(row, col + 1, r[1])
        row += 1
        a += 1

    chart.add_series({
        'values': '=Unrealized profit!$B$1:$B$' + str(a),
        'categories': '=Unrealized profit!$A$1:$A$' + str(a)
    })
    chart.set_size({'width': 720, 'height': 570})
    print("Finishing writing unrealized profit report at ", datetime.now())

    return chart

def showPlot(totalCumulativeChart):
    def format_date(x, pos=None):
        thisind = np.clip(int(x + 0.5), 0, N - 1)
        return totalCumulativeChart.index[thisind].strftime('%b %d %Y')

    N = len(totalCumulativeChart)
    ind = np.arange(N)
    fig, ax = plt.subplots()
    ax.plot(ind, totalCumulativeChart)
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_date))
    fig.autofmt_xdate()

    ax.yaxis.grid()
    plt.xticks(np.arange(min(ind), max(ind), 15))
    plt.show()

main()
