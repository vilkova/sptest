import Quandl as q
from numpy import *
import matplotlib.pyplot as plt
from openpyxl.reader.excel import load_workbook
import os
import sys
import re
import pickle
from datetime import datetime
import matplotlib.ticker as ticker
import numpy as np
import pandas as pd
import math
import xlsxwriter

AUTH_TOKEN = 'e6FuWkfWH9qypKzJz6sR'
CACHE_DIR = "cache-data/"
REPORTS_DIR = "reports/"

def main():
    if not os.path.exists(CACHE_DIR):
        os.makedirs(CACHE_DIR)
    if not os.path.exists(REPORTS_DIR):
        os.makedirs(REPORTS_DIR)
    table = retrieveTableFromExcel()
    spread1Delta = getSpreadDelta(table[1])
    if len(table) == 2:
        totalSpreadDelta = spread1Delta
    else:
        spread2Delta = getSpreadDelta(table[2])
        totalSpreadDelta = spread1Delta.add(spread2Delta, fill_value=0)
        for i in range(0, 3):
            del table[0]
        for row in table:
            totalSpreadDelta = totalSpreadDelta.add(getSpreadDelta(row), fill_value=0)
    convertDeltaAndShowPlot(totalSpreadDelta)

def retrieveTableFromExcel():
    table = loadSpreadMatrix(sys.argv[1])
    rowsWithOneContract = []
    if (len(sys.argv) == 5):
        contract = sys.argv[4]
        rowsWithOneContract.append(table[0])
        for row in table:
            if row[0].decode('utf-8') == contract:
                rowsWithOneContract.append(row)
        table = rowsWithOneContract
    return table

def loadSpreadMatrix(filename):
    wb = load_workbook(filename)
    table = []
    for sheet in wb.worksheets:
        for row in sheet.rows:
            table_row = []
            for cell in row:
                value = cell.value
                if value is None:
                    value = ''
                if not isinstance(value, str):
                    value = str(value)
                value = value.encode('utf8')
                table_row.insert(len(table_row), value)
            table.insert(len(table), table_row)
    return table

def getSpreadDelta(row):
    if len(sys.argv) == 2:
        years = [2000]
    else:
        years = range(int(sys.argv[2]), int(sys.argv[3]) + 1)
    spread = fetchSpread(row[0].decode("utf-8"), row[1].decode("utf-8"), row[2].decode("utf-8"),
                         int(row[5][:4].decode("utf-8")), int(row[6][:4].decode("utf-8")), int(row[3].decode("utf-8")),
                         int(row[4].decode("utf-8")), row[5].decode("utf-8"), row[6].decode("utf-8"),
                         int(row[7].decode("utf-8")), True, years)
    return convertSpreadSeriesToDelta(spread)

def checkIfCached(filename):
    fileNames = os.listdir(CACHE_DIR)
    for fileName in fileNames:
        if fileName == filename:
            return True
    return False

def readCacheFromFile(filename):
    cacheFile = open(CACHE_DIR + filename, "rb")
    cache = pickle.load(cacheFile)
    cacheFile.close()
    return cache

def fetchSpread(CONTRACT, M1, M2, ST_YEAR, END_YEAR, CONT_YEAR1, CONT_YEAR2, ST_DATE, END_DATE, BUCK_PRICE,
                STARTFROMZERO, years):
    startdate = datetime.strptime(ST_DATE, '%Y-%m-%d %H:%M:%S')
    enddate = datetime.strptime(END_DATE, '%Y-%m-%d %H:%M:%S')
    totalSpread = pd.Series()
    lastValue = 0
    for i in years:
        year = str(i)
        price = str(BUCK_PRICE)
        filename = CONTRACT + M1 + M2 + year + ST_DATE + END_DATE + price
        filename = re.sub('[/ ]', '_', filename)
        filename = re.sub('[:]', '.', filename)
        cont1 = str(CONTRACT) + str(M1) + str(i + CONT_YEAR1)
        cont2 = str(CONTRACT) + str(M2) + str(i + CONT_YEAR2)
        print('==============')
        print("contract1: " + cont1)
        print("contract2: " + cont2)
        startDate = startdate.replace(year=ST_YEAR - 2000 + i)
        endDate = enddate.replace(year=END_YEAR - 2000 + i)
        print('==============')
        print('Trim start: ', startDate.strftime('%Y-%m-%d'))
        print('Trim end: ', endDate.strftime('%Y-%m-%d'))
        print('==============')
        if not checkIfCached(filename):
            data1 = q.get(cont1, authtoken=AUTH_TOKEN, trim_start=startDate, trim_end=endDate)
            data2 = q.get(cont2, authtoken=AUTH_TOKEN, trim_start=startDate, trim_end=endDate)
            spread = (data1 - data2).Settle * BUCK_PRICE
            if spread.size == 0:
                print('!!!!!!!!!!!!*****WARNING****!!!!!!!!!!!!')
                print('No data available for contracts %s, %s. Skiping period from %s to %s.' % (
                    cont1, cont2, startDate.strftime('%Y-%m-%d'), endDate.strftime('%Y-%m-%d')))
                print('!!!!!!!!!!!!*****WARNING****!!!!!!!!!!!!')
                continue
            else:
                if math.isnan(spread[0]):
                    spread = spread.fillna(method='bfill')
                #replace NaN value with a previous one
                spread = spread.fillna(method='pad')

                #remove row with NAN value
                # spread = spread.dropna()
                writeCacheToFile(filename, spread)
        else:
            print("Loading cached data from file: %s !" % filename)
            cache = readCacheFromFile(filename)
            spread = cache
        if STARTFROMZERO:
            delta = lastValue - spread[0]
            spread = spread + delta
            totalSpread = totalSpread.append(spread)
            lastValue = totalSpread[-1]
    if totalSpread.size == 0:
        sys.exit(-1)
    return totalSpread

def writeCacheToFile(filename, spread):
    try:
        cacheFile = open(CACHE_DIR + filename, 'wb')
        pickle.dump(spread, cacheFile)
        cacheFile.close()
    except IOError:
        print('Error: can\'t write data to %s' % (CACHE_DIR + filename))

def convertSpreadSeriesToDelta(DATA):
    DATADELTA = DATA.copy(True)
    previ = DATA.index[0]
    for i in DATA.index:
        if DATA.index[0] != i:
            DATADELTA[i] = DATA.ix[i] - DATA.ix[previ]
            previ = i
    return DATADELTA

def convertDeltaAndShowPlot(totalSpreadDelta):
    totalCumulativeChart = convertDeltaSeriesToCumulativeGraph(totalSpreadDelta)
    dd = getDrawdowns(totalCumulativeChart)
    drawdownArray = foo(totalCumulativeChart)
    saveSortedDDArray(drawdownArray, dd)
    print('================')
    print('Maximum drawdowns: \n', sorted(dd, key=lambda x: x)[-5:], '\n')
    print('================')
    saveChartDataInFile(totalCumulativeChart)
    print("Total Cumulative Chart:")
    print(totalCumulativeChart.astype(int))
    showPlot(totalCumulativeChart)

def foo(chart):
    res = pd.Series()
    flag = False
    prev = 0
    ind = chart.index[0]
    res.set_value(ind, prev)
    for i in range(1, len(chart)):
        if flag:
            if chart[i] > prev:
                prev = chart[i]
                ind = chart.index[i]
                res.set_value(ind, 0)
                flag = False
            else:
                ind = chart.index[i]
                res.set_value(ind, chart[i] - prev)
        else:
            if chart[i] > prev:
                prev = chart[i]
                ind = chart.index[i]
                res.set_value(ind, 0)
            else:
                ind = chart.index[i]
                res.set_value(ind, chart[i] - prev)
                flag = True
    return res

def getDrawdowns(chart):
    filteredDDs = getMaxDrawdowns(chart)
    starts = []
    ends = []
    start = chart[1]
    for i in range(1, len(chart)):
        if chart[i] > start and chart[i] > chart[i-1]:
            if chart[i] > chart[i+1]:
                start = chart[i]
                startDate = chart.index[i]
                starts.append(startDate)
            if chart[i] >= start:
                    endDate = chart.index[i]
                    ends.append(endDate)
    return mix_dd(filteredDDs, ends)

def mix_dd(filteredDDs, ends):
    dates = []
    for i in range(0, len(filteredDDs)):
        startDate = filteredDDs[i][0]
        for j in range(1, len(ends)):
            if filteredDDs[i][0] < ends[j]:
                endDate = ends[j]
                dates.append((startDate, endDate, filteredDDs[i][2]))
                break
    return dates

def getMaxDrawdowns(totalCumulativeChart):
    maxValue = 0
    drawdownArray = []
    keyLeft = totalCumulativeChart.index[0]
    keyRight = totalCumulativeChart.index[0]
    for i in range(1, len(totalCumulativeChart) - 1):
        if totalCumulativeChart[i] > totalCumulativeChart[i - 1] and totalCumulativeChart[i] >= \
                totalCumulativeChart[i + 1]:
            if totalCumulativeChart[i] > maxValue:
                maxValue = totalCumulativeChart[i]
                keyRight = totalCumulativeChart.index[i]
        if totalCumulativeChart[i] < totalCumulativeChart[i - 1] and totalCumulativeChart[i] < totalCumulativeChart[
                    i + 1]:
            keyLeft = totalCumulativeChart.index[i]
            drawdownArray.append((keyRight, keyLeft, totalCumulativeChart[keyRight], totalCumulativeChart[keyLeft]))
    return filterDrawdowns(drawdownArray)

def filterDrawdowns(dd):
    keyMax = dd[0][0]
    keyMin = dd[0][1]
    maxValue = dd[0][2]
    minValue = dd[0][3]
    drawndowns = []
    for i in range(1, len(dd)):
        if dd[i][2] > maxValue:
            drawndowns.append((keyMax, keyMin, maxValue - minValue))
            keyMax = dd[i][0]
            keyMin = dd[i][1]
            maxValue = dd[i][2]
            minValue = dd[i][3]
        elif minValue > dd[i][3]:
            keyMin = dd[i][1]
            minValue = dd[i][3]
    return drawndowns

def convertDeltaSeriesToCumulativeGraph(DATA):
    GRAPHDATA = DATA.copy(True)
    prev_date = DATA.index[0]
    for i in range(1, len(DATA.index)):
        date = DATA.index[i]
        GRAPHDATA.ix[date] = GRAPHDATA.ix[prev_date] + DATA.ix[date]
        prev_date = date
    return GRAPHDATA

def saveSortedDDArray(sortedDDArray, dd):
    workbook = xlsxwriter.Workbook(REPORTS_DIR + 'drawdown_array.xlsx')
    worksheet = workbook.add_worksheet('Maximum drawdowns')
    worksheet1 = workbook.add_worksheet('All drawdowns')
    worksheet.set_column('A:B', 10)
    worksheet1.set_column('A:B', 10)
    worksheet1.set_column('M:O', 10)
    chart = getChartWithMaximumDrowdowns(workbook, worksheet, dd)
    chart1 = getChartWithAllDrawdowns(workbook, worksheet1, sortedDDArray, dd)
    worksheet.insert_chart('D1', chart)
    worksheet1.insert_chart('C1', chart1)
    workbook.close()

def getChartWithMaximumDrowdowns(workbook, worksheet, sortedDDArray):
    firstDate = []
    secondDate = []
    delta = []
    sortedDDArray = sorted(sortedDDArray, key=lambda x: x[2])[-5:]
    chart = workbook.add_chart({'type': 'column'})
    for i in range(0, len(sortedDDArray)):
        firstDate.append(sortedDDArray[i][0])
        secondDate.append(sortedDDArray[i][1])
        delta.append(sortedDDArray[i][2])
    col = 0
    row = 0
    for date1 in firstDate:
        dateLeft = datetime.strftime(date1, '%Y-%m-%d')
        worksheet.write_string(row, col, dateLeft)
        row += 1
    row = 0
    for date2 in secondDate:
        dateRight = datetime.strftime(date2, '%Y-%m-%d')
        worksheet.write_string(row, col + 1, dateRight)
        row += 1
    row = 0
    for d in delta:
        worksheet.write_number(row, col + 2, int(-d))
        row += 1
    chart.add_series({
        'values': '=Maximum drawdowns!$C$1:$C$5',
        'categories': '=Maximum drawdowns!$A$1:$A$5'
    })
    chart.set_size({'width': 600, 'height': 470})
    return chart

def getChartWithAllDrawdowns(workbook, worksheet1, sortedDDArray, dd):
    chart1 = workbook.add_chart({'type': 'area'})
    firstDate = []
    secondDate = []
    delta = []
    for i in range(0, len(dd)):
        firstDate.append(dd[i][0])
        secondDate.append(dd[i][1])
        delta.append(dd[i][2])
    writeArrayOFAllDDs(firstDate, secondDate, delta, worksheet1)
    col = 0
    row = 0
    a = 0
    c = 0
    for date1 in (sortedDDArray.index):
        dateLeft = datetime.strftime(date1, '%Y-%m-%d')
        worksheet1.write_string(row, col, dateLeft)
        row += 1
        a += 1
    # row = 0
    # for date2 in secondDate:
    #     dateRight = datetime.strftime(date2, '%Y-%m-%d')
    #     worksheet1.write_string(row, col + 1, dateRight)
    #     row += 1
    row = 0
    for d in (sortedDDArray):
        worksheet1.write_number(row, col + 1, int(d))
        row += 1
        c += 1
    chart1.add_series({
        'values': '=All drawdowns!$B$1:$B$' + str(c),
        'categories': '=All drawdowns!$A$1:$A$' + str(a)
    })
    chart1.set_size({'width': 600, 'height': 470})
    return chart1

def writeArrayOFAllDDs(firstDate, secondDate, delta, worksheet1):
    row = 0
    col = 12

    for date1 in firstDate:
        dateLeft = datetime.strftime(date1, '%Y-%m-%d')
        worksheet1.write_string(row, col, dateLeft)
        row += 1
    row = 0
    for date2 in secondDate:
        dateRight = datetime.strftime(date2, '%Y-%m-%d')
        worksheet1.write_string(row, col+1, dateRight)
        row += 1
    row = 0
    for d in delta:
        worksheet1.write_number(row, col + 2, int(-d))
        row += 1

def saveChartDataInFile(totalCumulativeChart):
    workbook = xlsxwriter.Workbook(REPORTS_DIR + 'chart_array.xlsx')
    worksheet = workbook.add_worksheet()
    chart = workbook.add_chart({'type': 'line'})
    a = 0
    b = 0
    row = 0
    col = 0
    for index in (totalCumulativeChart.index):
        date = datetime.strftime(index, '%Y-%m-%d')
        worksheet.write_string(row, col, date)
        a += 1
        row += 1
    row = 0
    for value in (totalCumulativeChart):
        worksheet.write_number(row, col + 1, int(value))
        b += 1
        row += 1
    chart.set_x_axis({
        'date_axis': True
    })
    chart.add_series({
        'values': '=Sheet1!$B$1:$B$' + str(b),
        'categories': '=Sheet1!$A$1:$A$' + str(a)
    })
    chart.set_y_axis({
        'major_gridlines': {
            'visible': True,
            'line': {'width': 1.25, 'dash_type': 'dash'}
        }
    })
    chart.set_size({'width': 720, 'height': 570})
    worksheet.insert_chart('C1', chart)
    workbook.close()

def showPlot(totalCumulativeChart):
    def format_date(x, pos=None):
        thisind = np.clip(int(x + 0.5), 0, N - 1)
        return totalCumulativeChart.index[thisind].strftime('%b %d %Y')

    N = len(totalCumulativeChart)
    ind = np.arange(N)

    #shows plot with empty data intervals    
    # fig, ax = plt.subplots()
    # ax.plot(totalCumulativeChart.index, totalCumulativeChart)
    # fig.autofmt_xdate()

    #shows plot without empty data intervals
    fig, ax = plt.subplots()
    ax.plot(ind, totalCumulativeChart)
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_date))
    fig.autofmt_xdate()

    ax.yaxis.grid()
    plt.xticks(np.arange(min(ind), max(ind), 15))
    plt.show()

main()
