import Quandl as q
from numpy import *
import matplotlib.pyplot as plt
from openpyxl.reader.excel import load_workbook
import os
import sys
import re
import pickle
from datetime import datetime
import matplotlib.ticker as ticker
import numpy as np
import pandas as pd
import math
import xlsxwriter

AUTH_TOKEN = 'e6FuWkfWH9qypKzJz6sR'
CACHE_DIR = "cache-data/"
B = sys.argv[4]

CURRENT_YEAR = datetime.now().year

def main():
    result = []
    if not os.path.exists(CACHE_DIR):
        os.makedirs(CACHE_DIR)
    table = retrieveTableFromExcel()
    spread1Delta = getSpreadDelta(table[1])
    result.append(countStdev(spread1Delta))
    if len(table) == 2:
        totalSpreadDelta = spread1Delta
    else:
        spread2Delta = getSpreadDelta(table[2])
        result.append(countStdev(spread2Delta))
        totalSpreadDelta = spread1Delta[1].add(spread2Delta[1], fill_value=0)
        for i in range(0, 3):
            del table[0]
        for row in table:
            spread = getSpreadDelta(row)
            result.append(countStdev(spread))
            totalSpreadDelta = totalSpreadDelta.add(spread[1], fill_value=0)
    totalCumulativeChart = convertDeltaSeriesToCumulativeGraph(totalSpreadDelta)
    print("Total Cumulative Chart:")
    print(totalCumulativeChart.astype(int))
    saveStdevs(result)
    showPlot(totalCumulativeChart)

def retrieveTableFromExcel():
    table = loadSpreadMatrix(sys.argv[1])
    rowsWithOneContract = []
    if (len(sys.argv) == 6):
        contract = sys.argv[5]
        rowsWithOneContract.append(table[0])
        for row in table:
            if row[0].decode('utf-8') == contract:
                rowsWithOneContract.append(row)
        table = rowsWithOneContract
    return table

def loadSpreadMatrix(filename):
    wb = load_workbook(filename)
    table = []
    for sheet in wb.worksheets:
        for row in sheet.rows:
            table_row = []
            for cell in row:
                value = cell.value
                if value is None:
                    value = ''
                if not isinstance(value, str):
                    value = str(value)
                value = value.encode('utf8')
                table_row.insert(len(table_row), value)
            table.insert(len(table), table_row)
    return table

def getSpreadDelta(row):
    if len(sys.argv) == 2:
        years = [2000]
    else:
        years = range(int(sys.argv[2]), int(sys.argv[3]) + 1)
    spread = fetchSpread(row[0].decode("utf-8"), row[1].decode("utf-8"), row[2].decode("utf-8"),
                         int(row[5][:4].decode("utf-8")), int(row[6][:4].decode("utf-8")), int(row[3].decode("utf-8")),
                         int(row[4].decode("utf-8")), row[5].decode("utf-8"), row[6].decode("utf-8"),
                         int(row[7].decode("utf-8")), True, years)
    spreadSeries = convertSpreadSeriesToDelta(spread[0])
    return (spread[1], spreadSeries, spread[2], spread[0], spread[3])

def fetchSpread(CONTRACT, M1, M2, ST_YEAR, END_YEAR, CONT_YEAR1, CONT_YEAR2, ST_DATE, END_DATE, BUCK_PRICE,
                STARTFROMZERO, years):
    startdate = datetime.strptime(ST_DATE, '%Y-%m-%d %H:%M:%S')
    enddate = datetime.strptime(END_DATE, '%Y-%m-%d %H:%M:%S')
    totalSpread = pd.Series()
    lastValue = 0
    filteredTotalSpread = pd.Series()
    spreadForStdevs = []
    daltasArray = []
    for i in years:
        year = str(i)
        price = str(BUCK_PRICE)
        filename = CONTRACT + M1 + M2 + year + ST_DATE + END_DATE + price
        filename = re.sub('[/ ]', '_', filename)
        filename = re.sub('[:]', '.', filename)
        cont1 = str(CONTRACT) + str(M1) + str(i + CONT_YEAR1)
        cont2 = str(CONTRACT) + str(M2) + str(i + CONT_YEAR2)
        print('==============')
        print("contract1: " + cont1)
        print("contract2: " + cont2)
        startDate = startdate.replace(year=ST_YEAR - 2000 + i)
        endDate = enddate.replace(year=END_YEAR - 2000 + i)
        print('==============')
        print('Trim start: ', startDate.strftime('%Y-%m-%d'))
        print('Trim end: ', endDate.strftime('%Y-%m-%d'))
        print('==============')
        if not checkIfCached(filename):
            data1 = q.get(cont1, authtoken=AUTH_TOKEN, trim_start=startDate, trim_end=endDate)
            data2 = q.get(cont2, authtoken=AUTH_TOKEN, trim_start=startDate, trim_end=endDate)
            spread = (data1 - data2).Settle * BUCK_PRICE
            if spread.size == 0:
                print('!!!!!!!!!!!!*****WARNING****!!!!!!!!!!!!')
                print('No data available for contracts %s, %s. Skiping period from %s to %s.' % (
                    cont1, cont2, startDate.strftime('%Y-%m-%d'), endDate.strftime('%Y-%m-%d')))
                print('!!!!!!!!!!!!*****WARNING****!!!!!!!!!!!!')
                continue
            else:
                if math.isnan(spread[0]):
                    spread = spread.fillna(method='bfill')
                #replace NaN value with a previous one
                spread = spread.fillna(method='pad')

                #remove row with NAN value
                # spread = spread.dropna()
                writeCacheToFile(filename, spread)
        else:
            print("Loading cached data from file: %s !" % filename)
            cache = readCacheFromFile(filename)
            spread = cache
        if STARTFROMZERO:
            deltas = convertSpreadSeriesToDelta(spread - spread[0])
            daltasArray.append(deltas)
            spreadForMeanReport = convertDeltaSeriesToCumulativeGraph(deltas)
            delta = lastValue - spread[0]
            spread = spread + delta
            spreadForStdevs.append(spreadForMeanReport[-1])
            totalSpread = totalSpread.append(spread)
            lastValue = totalSpread[-1]
    for i in range(0, len(totalSpread)):
        if totalSpread.index[i].year <= int(sys.argv[3]):
            filteredTotalSpread.set_value(totalSpread.index[i], totalSpread[i])
    if totalSpread.size == 0:
        sys.exit(-1)
    deltas = getDeltas(filteredTotalSpread)
    return (filteredTotalSpread, filename, spreadForStdevs, deltas)

def getDeltas(spread):
    deltas = []
    for i in range(1, len(spread)):
        prev = spread[i-1]
        next = spread[i]
        delta = next - prev
        deltas.append(delta)
        i += 1
    return deltas

def checkIfCached(filename):
    fileNames = os.listdir(CACHE_DIR)
    for fileName in fileNames:
        if fileName == filename:
            return True
    return False

def writeCacheToFile(filename, spread):
    try:
        cacheFile = open(CACHE_DIR + filename, 'wb')
        pickle.dump(spread, cacheFile)
        cacheFile.close()
    except IOError:
        print('Error: can\'t write data to %s' % (CACHE_DIR + filename))

def readCacheFromFile(filename):
    cacheFile = open(CACHE_DIR + filename, "rb")
    cache = pickle.load(cacheFile)
    cacheFile.close()
    return cache

def convertSpreadSeriesToDelta(DATA):
    DATADELTA = DATA.copy(True)
    previ = DATA.index[0]
    for i in DATA.index:
        if DATA.index[0] != i:
            DATADELTA[i] = DATA.ix[i] - DATA.ix[previ]
            previ = i
    return DATADELTA

def convertDeltaSeriesToCumulativeGraph(DATA):
    GRAPHDATA = DATA.copy(True)
    prev_date = DATA.index[0]
    for i in range(1, len(DATA.index)):
        date = DATA.index[i]
        GRAPHDATA.ix[date] = GRAPHDATA.ix[prev_date] + DATA.ix[date]
        prev_date = date
    return GRAPHDATA


def countStdev(spread):

    def stdev(x):
        return sqrt(sum((x - mean(x))**2)/(len(x)-1)) if len(x) > 1 else sqrt(sum((x - mean(x))**2)/len(x))
    contract = spread[0]
    series = spread[2]
    dailySeries = spread[3]
    deltas = spread[4]
    st_dev = stdev(series)
    avg = mean(series)
    daily_st_dev = stdev(dailySeries)
    deltas_avg = mean(deltas)
    deltas_stdev = stdev(deltas)

    return contract[:12], st_dev, avg, daily_st_dev, deltas_avg, deltas_stdev


def saveStdevs(list):

    workbook = xlsxwriter.Workbook('reports-2.xlsx')
    worksheet = workbook.add_worksheet("Coefficients")
    worksheet.set_column('A:A', 15)
    worksheet.write_string(0, 0, "Deal")
    worksheet.write_string(0, 1, "StDev(years)")
    worksheet.write_string(0, 2, "Max StDev/Stdev")
    worksheet.write_string(0, 3, "Coeft1")
    worksheet.write_string(0, 4, "Average")
    worksheet.write_string(0, 5, "Max Avg/Avg")
    worksheet.write_string(0, 6, "Coeft2")
    worksheet.write_string(0, 7, "StDev(days)")
    worksheet.write_string(0, 8, "Max StDev/Stdev")
    worksheet.write_string(0, 9, "Coeft3")

    worksheet.write_string(0, 10, "Average(deltas)")
    worksheet.write_string(0, 11, "MaxAvg/Avg")
    worksheet.write_string(0, 12, "Coeft4")
    worksheet.write_string(0, 13, "StDev(deltas)")
    worksheet.write_string(0, 14, "MaxStDev/Stdev")
    worksheet.write_string(0, 15, "Coeft5")
    col = 0
    row = 1
    max_dev = max(list,key=lambda x:x[1])[1]
    max_avg = max(list,key=lambda x:x[2])[2]
    max_daily_dev = max(list,key=lambda x:x[3])[3]
    max_deltas_avg = max(list,key=lambda x:x[4])[4]
    max_deltas_stdev = max(list,key=lambda x:x[5])[5]
    for element in list:
        st_dev = element[1]
        name = element[0]
        avg = element[2]
        daily_st_dev = element[3]
        delta_avg = element[4]
        delta_stdev = element[5]
        worksheet.write_string(row, col, str(name))

        value1 = max_dev/st_dev
        worksheet.write_number(row, col+1, round(st_dev))
        worksheet.write_number(row, col+2, round(value1))
        coeft1 = value1/int(B)
        worksheet.write_number(row, col+3, 1) if coeft1 <1 else worksheet.write_number(row, col+3, round(coeft1))

        value2 = max_avg/avg
        worksheet.write_number(row, col+4, round(avg))
        worksheet.write_number(row, col+5, round(value2))
        coeft2 = value2/int(B)
        worksheet.write_number(row, col+6, 1) if coeft2 <1 else worksheet.write_number(row, col+6, round(coeft2))

        value3 = max_daily_dev/daily_st_dev
        worksheet.write_number(row, col+7, round(daily_st_dev))
        worksheet.write_number(row, col+8, round(value3))
        coeft3 = value3/int(B)
        worksheet.write_number(row, col+9, 1) if coeft3 <1 else worksheet.write_number(row, col+9, round(coeft3))

        value4 = max_deltas_avg/delta_avg
        worksheet.write_number(row, col+10, round(delta_avg))
        worksheet.write_number(row, col+11, round(value4))
        coeft4 = value4/int(B)
        worksheet.write_number(row, col+12, 1) if coeft4 <1 else worksheet.write_number(row, col+12, round(coeft4))

        value5 = max_deltas_stdev/delta_stdev
        worksheet.write_number(row, col+13, round(delta_stdev))
        worksheet.write_number(row, col+14, round(value5))
        coeft5 = value5/int(B)
        worksheet.write_number(row, col+15, 1) if coeft5 <1 else worksheet.write_number(row, col+15, round(coeft5))

        row += 1


def showPlot(totalCumulativeChart):
    def format_date(x, pos=None):
        thisind = np.clip(int(x + 0.5), 0, N - 1)
        return totalCumulativeChart.index[thisind].strftime('%b %d %Y')

    N = len(totalCumulativeChart)
    ind = np.arange(N)

    #shows plot with empty data intervals
    # fig, ax = plt.subplots()
    # ax.plot(totalCumulativeChart.index, totalCumulativeChart)
    # fig.autofmt_xdate()

    #shows plot without empty data intervals
    fig, ax = plt.subplots()
    ax.plot(ind, totalCumulativeChart)
    ax.xaxis.set_major_formatter(ticker.FuncFormatter(format_date))
    fig.autofmt_xdate()

    ax.yaxis.grid()
    plt.xticks(np.arange(min(ind), max(ind), 15))
    plt.show()



main()
